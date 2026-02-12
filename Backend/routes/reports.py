from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from models import db, Faculty, Batch, FeedbackResponse
from utils.decorators import get_current_user, role_required
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import os
import io

reports_bp = Blueprint('reports', __name__)

PARAMETERS = [
    'Knowledge of the subject',
    'Coming well prepared for the class',
    'Giving clear explanations',
    'Command of language',
    'Clear and audible voice',
    'Holding the attention of students',
    'More matter than textbooks',
    'Clear doubts capability',
    'Encourage participation',
    'Appreciating students',
    'Help outside class',
    'Return papers on time',
    'Punctuality',
    'Coverage of syllabus',
    'Impartial'
]

@reports_bp.route('/faculty/<int:faculty_id>/excel', methods=['GET'])
@jwt_required()
def generate_faculty_excel(faculty_id):
    """Generate Excel report for a faculty member"""
    try:
        current_user = get_current_user()
        faculty = Faculty.query.get(faculty_id)
        
        if not faculty:
            return jsonify({'error': 'Faculty not found'}), 404
        
        # Check permissions
        if current_user.role == 'hod':
            if faculty.college != current_user.college or faculty.department != current_user.department:
                return jsonify({'error': 'Access denied'}), 403
        
        # Get all feedback for this faculty
        feedbacks = FeedbackResponse.query.filter_by(faculty_id=faculty_id).all()
        
        if not feedbacks:
            return jsonify({'error': 'No feedback data available'}), 404
        
        # Prepare data for Excel
        data = []
        for feedback in feedbacks:
            row = {
                'Submission Date': feedback.submitted_at.strftime('%Y-%m-%d %H:%M'),
                'Batch ID': feedback.batch_id,
                'Slot': feedback.slot_number,
            }
            
            # Add all parameter ratings
            for i in range(1, 16):
                row[PARAMETERS[i-1]] = getattr(feedback, f'param{i}_rating')
            
            row['Average'] = round(feedback.get_average_rating(), 2)
            row['Comments'] = feedback.comments or 'No comments'
            
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data
            df.to_excel(writer, sheet_name='Feedback Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Feedback Data']
            
            # Style the header
            header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_data = {
                'Parameter': PARAMETERS + ['Overall Average'],
                'Average Rating': []
            }
            
            for i in range(1, 16):
                param_avg = sum(getattr(f, f'param{i}_rating') for f in feedbacks) / len(feedbacks)
                summary_data['Average Rating'].append(round(param_avg, 2))
            
            overall_avg = sum(f.get_average_rating() for f in feedbacks) / len(feedbacks)
            summary_data['Average Rating'].append(round(overall_avg, 2))
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Style summary sheet
            summary_sheet = writer.sheets['Summary']
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight overall average
            last_row = len(PARAMETERS) + 2
            for cell in summary_sheet[last_row]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        
        output.seek(0)
        
        filename = f"Faculty_Report_{faculty.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/faculty/<int:faculty_id>/pdf', methods=['GET'])
@jwt_required()
def generate_faculty_pdf(faculty_id):
    """Generate PDF report for a faculty member"""
    try:
        current_user = get_current_user()
        faculty = Faculty.query.get(faculty_id)
        
        if not faculty:
            return jsonify({'error': 'Faculty not found'}), 404
        
        # Check permissions
        if current_user.role == 'hod':
            if faculty.college != current_user.college or faculty.department != current_user.department:
                return jsonify({'error': 'Access denied'}), 403
        
        feedbacks = FeedbackResponse.query.filter_by(faculty_id=faculty_id).all()
        
        if not feedbacks:
            return jsonify({'error': 'No feedback data available'}), 404
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0066CC'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            spaceAfter=10,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph('Faculty Feedback Report', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Faculty Information
        faculty_info = [
            ['Faculty Name:', faculty.name],
            ['Subject:', faculty.subject],
            ['Subject Code:', faculty.code],
            ['Department:', faculty.department],
            ['College:', faculty.college],
            ['Year/Branch:', f'{faculty.year} - {faculty.branch}'],
            ['Section:', faculty.section],
            ['Report Generated:', datetime.now().strftime('%B %d, %Y')]
        ]
        
        info_table = Table(faculty_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary Statistics
        elements.append(Paragraph('Summary Statistics', heading_style))
        
        total_responses = len(feedbacks)
        slot1_count = len([f for f in feedbacks if f.slot_number == 1])
        slot2_count = len([f for f in feedbacks if f.slot_number == 2])
        overall_avg = sum(f.get_average_rating() for f in feedbacks) / len(feedbacks)
        satisfaction = round((overall_avg / 5) * 100, 1)
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Responses', str(total_responses)],
            ['Slot 1 Responses', str(slot1_count)],
            ['Slot 2 Responses', str(slot2_count)],
            ['Overall Average Rating', f'{round(overall_avg, 2)}/10.0'],
            ['Satisfaction Percentage', f'{satisfaction}%']
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Parameter-wise Analysis
        elements.append(Paragraph('Parameter-wise Average Ratings', heading_style))
        
        param_data = [['Parameter', 'Average Rating']]
        
        for i in range(1, 16):
            param_avg = sum(getattr(f, f'param{i}_rating') for f in feedbacks) / len(feedbacks)
            param_data.append([PARAMETERS[i-1], f'{round(param_avg, 2)}/10.0'])
        
        param_table = Table(param_data, colWidths=[4.5*inch, 1.5*inch])
        param_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(param_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"Faculty_Report_{faculty.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/batch/<batch_id>/excel', methods=['GET'])
@jwt_required()
def generate_batch_excel(batch_id):
    """Generate Excel report for entire batch"""
    try:
        current_user = get_current_user()
        batch = Batch.query.filter_by(batch_id=batch_id).first()
        
        if not batch:
            return jsonify({'error': 'Batch not found'}), 404
        
        # Check permissions
        if current_user.role == 'hod':
            if batch.college != current_user.college or batch.department != current_user.department:
                return jsonify({'error': 'Access denied'}), 403
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create sheet for each faculty
            for faculty in batch.faculty:
                feedbacks = FeedbackResponse.query.filter_by(
                    batch_id=batch_id,
                    faculty_id=faculty.id
                ).all()
                
                if feedbacks:
                    data = []
                    for feedback in feedbacks:
                        row = {'Submission Date': feedback.submitted_at.strftime('%Y-%m-%d %H:%M')}
                        
                        for i in range(1, 16):
                            row[PARAMETERS[i-1][:20]] = getattr(feedback, f'param{i}_rating')
                        
                        row['Average'] = round(feedback.get_average_rating(), 2)
                        data.append(row)
                    
                    df = pd.DataFrame(data)
                    sheet_name = faculty.name[:30]  # Excel sheet name limit
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        
        filename = f"Batch_Report_{batch_id[:30]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
